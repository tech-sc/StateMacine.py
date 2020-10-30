# -*- coding: utf-8 -*-
"""
Created on Sun Jul  5 16:38:14 2020

@author: teru
"""

import openpyxl


StateList = []
EventList = []
ProcList = []
StateTableList = []
NextStateList = []

def readStateMatrix_Excel():
	try:
		book = openpyxl.load_workbook('StateMachine.Base.xlsx')
		sheet = book['StateMatrix']
		print(type(sheet))
		return sheet
	except Exception as err:
		print('Excel file not found.' + str(err))

def writeCFile(str):
    print(str)

# State Name cell(D4:D5,E4:E5,...)
# code is D4:D5->E4:E5->F4:F5...
def writeStateEnum(sheet):
	writeCFile('typedef enum {')
	i = 0
	for code in sheet.iter_cols(min_row=4, min_col=4, max_row=5, values_only=True):
		if (code[0] == None and code[1] == None):
			break
		writeCFile('\t// ' + code[0])
		writeCFile('\t' + code[1] + ',')
		StateList.append(code[1])
		i += 1
	writeCFile('\tMAX_STATE_ID')
	writeCFile('} STATE_ID_e;\n')
	return i + 4


# Event Name cell(B6,B7,...)
# Analize: B6->B7->B8->B9...
def writeEventEnum(sheet):
	writeCFile('typedef enum {')
	i = 0
	for code in sheet.iter_rows(min_row=6, min_col=2, max_col=2, values_only=True):
		if i & 1 == 0:
			writeCFile('\t// ' + code[0])
		else:
			writeCFile('\t' + code[0] + ',')
			EventList.append(code[0])
		i += 1
	writeCFile('\tMAX_EVENT_ID')
	writeCFile('} EVENT_ID_e;\n')
	return i + 6

def writeStateTypedef():
	writeCFile('typedef struct {')
	writeCFile('\tEVENT_ID_e\t\tevent_id;')
	writeCFile('\tint32_t\t\t\t(*event_proc)(MSG_t *msg);')
	writeCFile('} STATE_TABLE_t;\n')

# Action cell(D6:n99)
# Analize: D6->D7->D8...E6->E7...
def writeStateTable(sheet):
	st = 0
	for cols in sheet.iter_cols(min_row=6, min_col=4, max_col=max_x-1):
		writeCFile('STATE_TABLE_t\t' + StateList[st] +'_Table[] = {')
		StateTableList.append(StateList[st] +'_Table')
		next = False
		i = 0
		for cell in cols:
			if i & 1 == 0:
				if cell.value != '無視':
					# writeCFile(f'{cell.row}Line: {cell.value}')
					writeCFile('\t{ ' + EventList[int(i/2)] + ',\t\t' + cell.value + ' },')
					ProcList.append(cell.value)
					next = True
			else:
				if next:
					NextStateList.append(cell.value)
					next = False
			i += 1
		writeCFile('\t{ NULL,NULL }')
		writeCFile('};\n')
		st += 1

def writeRootStateTable():
	writeCFile('STATE_TABLE_t\t*RootStateTable[MAX_STATE_ID] = {')
	for code in StateTableList:
		writeCFile(f'\t{code},')
	writeCFile('};\n')

def writeActionProtetype():
	writeCFile('//TODO: move to Top line')
	for code in ProcList:
		writeCFile(f'int32_t {code}( MSG_t *msg );')
	writeCFile('//TODO: end\n')

def writeActionProcedure():
	st = 0
	for code in ProcList:
		writeCFile(f'int32_t {code}( MSG_t *msg )')
		writeCFile('{\n\t// ' + NextStateList[st])
		writeCFile('}\n')
		st += 1
	writeCFile('//Please, append C source: StateMachine.func.c\n')


sheet = readStateMatrix_Excel()
max_x = writeStateEnum(sheet)
max_y = writeEventEnum(sheet)
writeStateTypedef()
writeStateTable(sheet)
writeRootStateTable()
writeActionProtetype()
writeActionProcedure()
